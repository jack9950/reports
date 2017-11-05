import sys
import openpyxl
from openpyxl.styles import Font, PatternFill
import time
from datetime import datetime
from MTD import get_calls_handled, get_pogo_sales, get_DEPP_sales, get_fcp_sales, get_HIVE_new_service, get_HIVE_renewals
from MTD import homeFolder, callsHandledReportLocation, pogoSalesReportLocation
from MTD import fcpReportLocation, DEPPreportLocation, hiveNewServiceReportLocation, hiveRenewalsReportLocation

#Cell Background and Font Styles (to be used to conditionally format cells)
below_goal_text = "9C0006"
below_goal_bg = "FFC7CE"
close_to_goal_text = "9C6500"
close_to_goal_bg = "FFEB9C"
at_or_above_goal_text = "006100"
at_or_above_goal_bg = "C6EFCE"

jaelesiaTeam = [2062001, 2062011, 2062020, 2062026, 2062036, 2062048, 2062053,
                2062054, 2062057, 2062062]
tekTeam = [2062067, 2062051, 2062035, 2062015, 2062040, 2062010, 2062042,
           2062024, 2062065, 2062060, 2062007]
antwonTeam = [2062039, 2062073, 2062074, 2062052, 2062058, 2062018, 2062049,
              2062076, 2062031, 2062044, 2062003, 2062032, 2062066]

jacksonTeam = [2062090, 2062081, 2062082, 2062083, 2062084, 2062085, 2062086, 2062089]

agentIDs = [2062004, 2062026, 2062043, 2062034, 2062053, 2062048, 2062042,
            2062011, 2062030, 2062045, 2062046, 2062016, 2062001, 2062036,
            2062039, 2062025, 2062041, 2062052, 2062037, 2062024, 2062049,
            2062031, 2062044, 2062003, 2062028, 2062022, 2062051, 2062021,
            2062035, 2062007, 2062020, 2062015, 2062040, 2062010, 2062018,
            2062054, 2062032, 2062033, 2062062, 2062070, 2062067, 2062058,
            2062056, 2062066, 2062057, 2062065, 2062060, 2062090, 2062081,
            2062082, 2062083, 2062084, 2062085, 2062086, 2062089]

jaelesiaTotalCallsHandled, tekTotalCallsHandled, antwonTotalCallsHandled, jacksonTotalCallsHandled, totalCallsHandled = 0, 0, 0, 0, 0
jaelesiaSalesCallsHandled, tekSalesCallsHandled, antwonSalesCallsHandled, jacksonSalesCallsHandled, totalSalesCallsHandled = 0, 0, 0, 0, 0
jaelesiaTotalSales, tekTotalSales, antwonTotalSales, jacksonTotalSales, totalSales = 0, 0, 0, 0, 0
jaelesiaFCPsales, tekFCPsales, antwonFCPsales, jacksonFCPsales, totalFCPSales= 0, 0, 0, 0, 0
jaelesiaNestSales, tekNestSales, antwonNestSales, jacksonNestSales, totalNestSales= 0, 0, 0, 0, 0
jaelesiaDEPPsales, tekDEPPsales, antwonDEPPsales, jacksonDEPPsales, totalDEPPsales= 0, 0, 0, 0, 0
jaelesiaHiveSales, tekHiveSales, antwonHiveSales, jacksonHiveSales, totalHiveSales = 0, 0, 0, 0, 0

supervisorIDs = {"aervin":2062007, "jnickerson":2062001, "tlevon": 2062007,
                 "jacksonn": 2062047, "jabram":2062017, "iqr_acollins":2062072,
                 "jmoore":2062023, "mayala":2062002}

#Open the template file for editing:
print("\nOpening template file for editing......\n")

template = openpyxl.load_workbook(homeFolder + 'MTD\\template_sales_MTD.xlsx')
template_sheets = template.get_sheet_names()
template_first_sheet = template.get_sheet_by_name(template_sheets[0])


#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
#Open the Bounce Hourly Sales Report from iQor,
#Retrieve data and add to summary tab of template excel file
#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------

#Get the calls handled for each agent

print("\nReading agent IDs and call counts.......\n")

#The format returned by get_calls_handled is a 2 dimensional array with each agent
#and their calls represented as: [agent ID, Calls Handled, Sales Calls Handled] in the return array

calls_handled = get_calls_handled(callsHandledReportLocation)

#Write out the call counts to the template file
print("\nWriting call counts to the template file.......\n")
for row in range(3, 60):
    agent_id_cell = "A" + str(row)
    agent_id = template_first_sheet[agent_id_cell].value
    #retrieve each agent ID from the template and check if it is the list.
    #if found write the calls handled and sales calls handled to the template file
    if(agent_id != None):
        for item in calls_handled: #check each nested list
            if agent_id in item and item[1] != 0: #if found and total calls > 0,
                                                  #write calls data to the template
                template_first_sheet["C"+str(row)].value = item[1] #Total Calls Handled
                template_first_sheet["D"+str(row)].value = item[2] #Sales Calls Handled

#Sum up the calls handled for each supervisor and for the whole of iQor
for item in calls_handled:
    agent_id = item[0]
    if agent_id in jaelesiaTeam:
        jaelesiaTotalCallsHandled += item[1]
        jaelesiaSalesCallsHandled += item[2]
        totalCallsHandled += item[1]
        totalSalesCallsHandled += item[2]
    if agent_id in tekTeam:
        tekTotalCallsHandled += item[1]
        tekSalesCallsHandled += item[2]
        totalCallsHandled += item[1]
        totalSalesCallsHandled += item[2]
    if agent_id in antwonTeam:
        antwonTotalCallsHandled += item[1]
        antwonSalesCallsHandled += item[2]
        totalCallsHandled += item[1]
        totalSalesCallsHandled += item[2]
    if agent_id in jacksonTeam:
        jacksonTotalCallsHandled += item[1]
        jacksonSalesCallsHandled += item[2]
        totalCallsHandled += item[1]
        totalSalesCallsHandled += item[2]

#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
#Gather up all the orders from the big bounce sales report and
#write them out to the template
#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------

print("\nGathering up all the orders from the big bounce sales report.......\n")

pogo_sales = get_pogo_sales(pogoSalesReportLocation)

#Team leads usually submit POGO orders with their text POGO ID rather than the numeric one
#Replace the team lead text POGO agent IDs with the numeric AVAYA IDs
for id in pogo_sales:
    if (type(id) == str):
        try:
            pogo_sales[pogo_sales.index(id)] = supervisorIDs[id]
        except:
            pass

#Write out the number of sales to the template
print("\nWriting out the number of sales to the template.......\n")
for i in range(3, 50):
    agent_id_cell = "A"+str(i)
    agent_id = template_first_sheet[agent_id_cell].value
    calls_handled_cell = "C"+str(i)
    calls_handled = template_first_sheet[calls_handled_cell].value
    if(agent_id != None and calls_handled != None):
        template_first_sheet["E"+str(i)].value = pogo_sales.count(agent_id)

#Sum up the POGO sales for each supervisor and for the whole of iQor
for agent_id in pogo_sales:
    if agent_id in jaelesiaTeam:
        jaelesiaTotalSales += 1
        totalSales += 1
    if agent_id in tekTeam:
        tekTotalSales += 1
        totalSales += 1
    if agent_id in antwonTeam:
        antwonTotalSales += 1
        totalSales += 1
    if agent_id in jacksonTeam:
        jacksonTotalSales += 1
        totalSales += 1

#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
#Gather up the DEPP sales from the Products report and
#write them out to the template
#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------

DEPP_sales = get_DEPP_sales(DEPPreportLocation)

for id in DEPP_sales:
    if (type(id) == str):
        try:
            DEPP_sales[DEPP_sales.index(id)] = supervisorIDs[id]
        except:
            pass

#Write out the products to the template
print("\nWriting the products to the template..........\n")
for i in range(3, 50):
    agent_id_cell = "A"+str(i)
    calls_handled_cell = "C"+str(i)
    agent_id = template_first_sheet[agent_id_cell].value
    calls_handled = template_first_sheet[calls_handled_cell].value
    if(agent_id != None and calls_handled != None):
        #template_first_sheet["H"+str(i)].value = nest_sales.count(agent_id)
        template_first_sheet["H"+str(i)].value = DEPP_sales.count(agent_id)

#Sum up the DEPP sales for each supervisor and for the whole of iQor
for agent_id in DEPP_sales:
    if agent_id in jaelesiaTeam:
        jaelesiaDEPPsales += 1
        totalDEPPsales += 1
    if agent_id in tekTeam:
        tekDEPPsales += 1
        totalDEPPsales += 1
    if agent_id in antwonTeam:
        antwonDEPPsales += 1
        totalDEPPsales += 1
    if agent_id in jacksonTeam:
        jacksonDEPPsales += 1
        totalDEPPsales += 1

#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
#Gather up the FCP sales from the FCP report and
#write them out to the template
#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------

print("\nOpening fcp report......\n")

fcp_sales = get_fcp_sales(fcpReportLocation)

#Write out the FCP sales to the template
print("\nWriting out the FCP sales to the template.......\n")
for i in range(3, 50):
    agent_id_cell = "A"+str(i)
    agent_id = template_first_sheet[agent_id_cell].value
    calls_handled_cell = "C"+str(i)
    calls_handled = template_first_sheet[calls_handled_cell].value
    if(agent_id != None and calls_handled != None):
        template_first_sheet["G"+str(i)].value = fcp_sales.count(agent_id)

#Sum up the fcp sales for each supervisor and for the whole of iQor
for agent_id in fcp_sales:
    if agent_id in jaelesiaTeam:
        jaelesiaFCPsales += 1
        totalFCPSales += 1
    if agent_id in tekTeam:
        tekFCPsales += 1
        totalFCPSales += 1
    if agent_id in antwonTeam:
        antwonFCPsales += 1
        totalFCPSales += 1
    if agent_id in jacksonTeam:
        jacksonFCPsales += 1
        totalFCPSales += 1

#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
#Gather up the HIVE sales from the Products report and HIVE Renewals report
#write them out to the template
#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------

hive_new_service_sales = get_HIVE_new_service(hiveNewServiceReportLocation)
hive_renewal_sales = get_HIVE_renewals(hiveRenewalsReportLocation)
# hive_renewal_sales = get_HIVE_renewals()

all_HIVE_sales = []

for sale in hive_new_service_sales:
    if sale not in all_HIVE_sales:
        all_HIVE_sales.append(sale)

for sale in hive_renewal_sales:
    if sale not in all_HIVE_sales:
        all_HIVE_sales.append(sale)

IDs_with_HIVE_sales = []

for sale in all_HIVE_sales:
    IDs_with_HIVE_sales.append(sale[0])

print("\nWriting out the HIVE sales to the template.......\n")
for i in range(3, 50):
    agent_id_cell = "A"+str(i)
    agent_id = template_first_sheet[agent_id_cell].value
    calls_handled_cell = "C"+str(i)
    calls_handled = template_first_sheet[calls_handled_cell].value
    if(agent_id != None and calls_handled != None):
        template_first_sheet["I"+str(i)].value = IDs_with_HIVE_sales.count(agent_id)

#-------------------------------------------------------------------------------
#Sum up the HIVE sales for each supervisor and for the whole of iQor
#-------------------------------------------------------------------------------
for agent_id in IDs_with_HIVE_sales:
    if agent_id in jaelesiaTeam:
        jaelesiaHiveSales += 1
        totalHiveSales += 1
    if agent_id in tekTeam:
        tekHiveSales += 1
        totalHiveSales += 1
    if agent_id in antwonTeam:
        antwonHiveSales += 1
        totalHiveSales += 1
    if agent_id in jacksonTeam:
        jacksonHiveSales += 1
        totalHiveSales += 1

print("jacksonTotalCallsHandled: ", jacksonTotalCallsHandled)
#-------------------------------------------------------------------------------
#Write out the agent close rates to the template
#-------------------------------------------------------------------------------
for i in range(3,50):
    try:
        closeRate = (template_first_sheet["e" + str(i)].value + template_first_sheet["g" + str(i)].value) / template_first_sheet["d" + str(i)].value
        template_first_sheet["f" + str(i)].value = closeRate
        closeRateCell = template_first_sheet["f" + str(i)]
        # print(closeRate)
        if closeRate < 0.4:
            closeRateCell.font = Font(name='Calibri', size=11, bold=True, color=below_goal_text)
            closeRateCell.fill = PatternFill("solid", fgColor=below_goal_bg)
        elif closeRate >= 0.5:
            closeRateCell.font = Font(name='Calibri', size=11, bold=True, color=at_or_above_goal_text)
            closeRateCell.fill = PatternFill("solid", fgColor=at_or_above_goal_bg)
        else:
            closeRateCell.font = Font(name='Calibri', size=11, bold=True, color=close_to_goal_text)
            closeRateCell.fill = PatternFill("solid", fgColor=close_to_goal_bg)
    except:
        pass

    #Write out the supervisor and iQor totals to the template
    #Add all the fonts and the conditional color formats
    if template_first_sheet["b" + str(i)].value == "JAELESIA MOORE Total":
        template_first_sheet["c" + str(i)].value = jaelesiaTotalCallsHandled
        template_first_sheet["d" + str(i)].value = jaelesiaSalesCallsHandled
        template_first_sheet["e" + str(i)].value = jaelesiaTotalSales
        template_first_sheet["g" + str(i)].value = jaelesiaFCPsales
        template_first_sheet["h" + str(i)].value = jaelesiaDEPPsales
        template_first_sheet["i" + str(i)].value = jaelesiaHiveSales
        try:
            closeRate = (template_first_sheet["e" + str(i)].value + template_first_sheet["g" + str(i)].value) / template_first_sheet["d" + str(i)].value
            template_first_sheet["f" + str(i)].value = closeRate
            closeRateCell = template_first_sheet["f" + str(i)]
        except:
            pass

        closeRateCell = template_first_sheet["f" + str(i)]
        if closeRate < 0.4:
            closeRateCell.font = Font(name='Calibri', size=13, bold=True, color=below_goal_text)
            closeRateCell.fill = PatternFill("solid", fgColor=below_goal_bg)
        elif closeRate >= 0.5:
            closeRateCell.font = Font(name='Calibri', size=13, bold=True, color=at_or_above_goal_text)
            closeRateCell.fill = PatternFill("solid", fgColor=at_or_above_goal_bg)
        else:
            closeRateCell.font = Font(name='Calibri', size=13, bold=True, color=close_to_goal_text)
            closeRateCell.fill = PatternFill("solid", fgColor=close_to_goal_bg)

    if template_first_sheet["b" + str(i)].value == "TEK LEVON Total":
        template_first_sheet["c" + str(i)].value = tekTotalCallsHandled
        template_first_sheet["d" + str(i)].value = tekSalesCallsHandled
        template_first_sheet["e" + str(i)].value = tekTotalSales
        template_first_sheet["g" + str(i)].value = tekFCPsales
        template_first_sheet["h" + str(i)].value = tekDEPPsales
        template_first_sheet["i" + str(i)].value = tekHiveSales
        try:
            closeRate = (template_first_sheet["e" + str(i)].value + template_first_sheet["g" + str(i)].value) / template_first_sheet["d" + str(i)].value
            template_first_sheet["f" + str(i)].value = closeRate
            closeRateCell = template_first_sheet["f" + str(i)]
        except:
            pass

        closeRateCell = template_first_sheet["f" + str(i)]

        if closeRate < 0.4:
            closeRateCell.font = Font(name='Calibri', size=13, bold=True, color=below_goal_text)
            closeRateCell.fill = PatternFill("solid", fgColor=below_goal_bg)
        elif closeRate >= 0.5:
            closeRateCell.font = Font(name='Calibri', size=13, bold=True, color=at_or_above_goal_text)
            closeRateCell.fill = PatternFill("solid", fgColor=at_or_above_goal_bg)
        else:
            closeRateCell.font = Font(name='Calibri', size=13, bold=True, color=close_to_goal_text)
            closeRateCell.fill = PatternFill("solid", fgColor=close_to_goal_bg)

    if template_first_sheet["b" + str(i)].value == "ANTWON COLLINS Total":
        template_first_sheet["c" + str(i)].value = antwonTotalCallsHandled
        template_first_sheet["d" + str(i)].value = antwonSalesCallsHandled
        template_first_sheet["e" + str(i)].value = antwonTotalSales
        template_first_sheet["g" + str(i)].value = antwonFCPsales
        template_first_sheet["h" + str(i)].value = antwonDEPPsales
        template_first_sheet["i" + str(i)].value = antwonHiveSales
        try:
            closeRate = (template_first_sheet["e" + str(i)].value + template_first_sheet["g" + str(i)].value) / template_first_sheet["d" + str(i)].value
            template_first_sheet["f" + str(i)].value = closeRate
            closeRateCell = template_first_sheet["f" + str(i)]
        except:
            pass

        closeRateCell = template_first_sheet["f" + str(i)]

        if closeRate < 0.4:
            closeRateCell.font = Font(name='Calibri', size=13, bold=True, color=below_goal_text)
            closeRateCell.fill = PatternFill("solid", fgColor=below_goal_bg)
        elif closeRate >= 0.5:
            closeRateCell.font = Font(name='Calibri', size=13, bold=True, color=at_or_above_goal_text)
            closeRateCell.fill = PatternFill("solid", fgColor=at_or_above_goal_bg)
        else:
            closeRateCell.font = Font(name='Calibri', size=13, bold=True, color=close_to_goal_text)
            closeRateCell.fill = PatternFill("solid", fgColor=close_to_goal_bg)

    if template_first_sheet["b" + str(i)].value == "JACKSON NDIHO Total":
        template_first_sheet["c" + str(i)].value = jacksonTotalCallsHandled

        template_first_sheet["d" + str(i)].value = jacksonSalesCallsHandled
        template_first_sheet["e" + str(i)].value = jacksonTotalSales
        template_first_sheet["g" + str(i)].value = jacksonFCPsales
        template_first_sheet["h" + str(i)].value = jacksonDEPPsales
        template_first_sheet["i" + str(i)].value = jacksonHiveSales
        try:
            closeRate = (template_first_sheet["e" + str(i)].value + template_first_sheet["g" + str(i)].value) / template_first_sheet["d" + str(i)].value
            template_first_sheet["f" + str(i)].value = closeRate
            closeRateCell = template_first_sheet["f" + str(i)]
        except:
            pass

        closeRateCell = template_first_sheet["f" + str(i)]
        if closeRate < 0.4:
            closeRateCell.font = Font(name='Calibri', size=13, bold=True, color=below_goal_text)
            closeRateCell.fill = PatternFill("solid", fgColor=below_goal_bg)
        elif closeRate >= 0.5:
            closeRateCell.font = Font(name='Calibri', size=13, bold=True, color=at_or_above_goal_text)
            closeRateCell.fill = PatternFill("solid", fgColor=at_or_above_goal_bg)
        else:
            closeRateCell.font = Font(name='Calibri', size=13, bold=True, color=close_to_goal_text)
            closeRateCell.fill = PatternFill("solid", fgColor=close_to_goal_bg)

    if template_first_sheet["b" + str(i)].value == "Grand Total":
        template_first_sheet["c" + str(i)].value = totalCallsHandled
        template_first_sheet["d" + str(i)].value = totalSalesCallsHandled
        template_first_sheet["e" + str(i)].value = totalSales
        template_first_sheet["g" + str(i)].value = totalFCPSales
        template_first_sheet["h" + str(i)].value = totalDEPPsales
        template_first_sheet["i" + str(i)].value = totalHiveSales
        try:
            closeRate = (template_first_sheet["e" + str(i)].value + template_first_sheet["g" + str(i)].value) / template_first_sheet["d" + str(i)].value
            template_first_sheet["f" + str(i)].value = closeRate
            closeRateCell = template_first_sheet["f" + str(i)]
        except:
            pass

        closeRateCell = template_first_sheet["f" + str(i)]

        if closeRate < 0.4:
            closeRateCell.font = Font(name='Calibri', size=13, bold=True, color=below_goal_text)
            closeRateCell.fill = PatternFill("solid", fgColor=below_goal_bg)
        elif closeRate >= 0.5:
            closeRateCell.font = Font(name='Calibri', size=13, bold=True, color=at_or_above_goal_text)
            closeRateCell.fill = PatternFill("solid", fgColor=at_or_above_goal_bg)
        else:
            closeRateCell.font = Font(name='Calibri', size=13, bold=True, color=close_to_goal_text)
            closeRateCell.fill = PatternFill("solid", fgColor=close_to_goal_bg)

#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
#We are done! - Save the template as final.xlsx
#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
print("\nSaving final template.......")

finalReportName = 'SalesReport_MTD'
currentDate = datetime.now().strftime("%m%d%Y")
currentTime = time.strftime("%I%M%S%p")

template.save(homeFolder + 'MTD\\' + finalReportName + "_" + currentDate + "_" + currentTime + ".xlsx")

print("\nDone.......\n")
